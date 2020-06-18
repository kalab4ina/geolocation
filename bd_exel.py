import sys
import os
import openpyxl
import csv
from geo_place import extract_lat_long_via_address


bd_open = openpyxl.load_workbook(filename = 'BD_main.xlsx',read_only=True)
bd_sheet = bd_open.sheetnames[0]
work_sheet =  bd_open[bd_sheet]

data = []
for i in range(3,6):
    data += [[work_sheet['F'+str(i)].value,
            work_sheet['B'+str(i)].value,
            work_sheet['D'+str(i)].value,
            work_sheet['E'+str(i)].value,
            extract_lat_long_via_address(work_sheet['B'+str(i)].value)]]

new_base = openpyxl.Workbook()
sheet = new_base.active
sheet.title = 'data'

#######
row = 1
sheet['A'+str(row)] = work_sheet['F'+str(1)].value
sheet['B'+str(row)] = work_sheet['B'+str(1)].value
sheet['C'+str(row)] = work_sheet['D'+str(1)].value
sheet['D'+str(row)] = work_sheet['E'+str(1)].value
sheet['E'+str(row)] = 'Долгота новая'
sheet['F'+str(row)] = 'Широта новая'
######
for item in data:
    row += 1
    sheet['A'+str(row)] = item[0]
    sheet['B'+str(row)] = item[1]
    sheet['C'+str(row)] = item[2]
    sheet['D'+str(row)] = item[3]
    sheet['E'+str(row)] = item[4][0]
    sheet['F'+str(row)] = item[4][1]


#####
filename = 'base_new.xlsx'
new_base.save(filename)
#####
os.chdir(sys.path[0])
os.system('start excel.exe "%s\\%s"' % (sys.path[0], filename,))


    
    
